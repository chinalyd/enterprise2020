import datetime
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook('courses.xlsx')
students_sheet = wb['students']
time_sheet = wb['time']

def combine():
    combine_sheet = wb.create_sheet(title = 'combine')#Add Sheet Title
    combine_sheet.append(['create courese', 'course name', 'number of learner', 'learn time'])#Add content to table
    for stu in students_sheet.values:
        if stu[2] != '学习人数':
            for time in time_sheet.values:
                if time[1] == stu[1]:
                    combine_sheet.append(list(stu) + [time[2]])#Add line to combine
    wb.save('courses.xlsx')

def split():
    combine_sheet = wb['combine']
    split_name = [] #save year record
    for item in combine_sheet.values:
        if item[0] != '创建时间' and type(item[0]) != str:
            split_name.append(item[0].strftime("%Y"))
    for name in set(split_name): #save each data
        wb_temp = Workbook()#create file
        wb_temp.remove(wb_temp.active)#delete saved sheet
        ws = wb_temp.create_sheet(title=name)
        ws.append(['创建时间', '课程名称', '学习人数', '学习时间'])
        for item_by_year in combine_sheet.values:
            if item_by_year[0] != '创建时间' and type(item_by_year[0]) != str:
                if item_by_year[0].strftime("%Y") == name:
                    ws.append(item_by_year)
        wb_temp.save('{}.xlsx'.format(name))

if __name__ == '__main__':
    combine()
    split()
    
